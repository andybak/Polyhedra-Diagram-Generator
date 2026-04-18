import csv
import math
import os
import sys
import subprocess
import platform
import itertools
from dataclasses import dataclass
sys.modules['cairocffi'] = None  # force svglib/reportlab to use pycairo instead of cairocffi
from svglib.svglib import svg2rlg
from reportlab.graphics import renderPM

dot_positions = {
    'V': [(0, 0), (1, 0), (1, 1), (0, 1)],
    'F': [(0.5, 0.5)],
    'F!': [(0.5, -0.2), (1.2, 0.5), (0.5, 1.2), (-0.2, 0.5)],
    'E': [(0.5, 0), (1, 0.5), (0.5, 1), (0, 0.5)],
    'vf': [(0.25, 0.25), (0.75, 0.25), (0.75, 0.75), (0.25, 0.75)],
    'vf!': [(0.25, -0.2), (0.75, -0.2), (1.2, 0.25), (1.2, 0.75), (0.75, 1.2), (0.25, 1.2), (-0.2, 0.75), (-0.2, 0.25)],
    've': [(0.25, 0), (0.75, 0), (1, 0.25), (1, 0.75), (0.75, 1), (0.25, 1), (0, 0.75), (0, 0.25)],
    've0': [(0.25, 0), (0.75, 0), (1, 0.25), (1, 0.75), (0.75, 1), (0.25, 1), (0, 0.75), (0, 0.25)],
    've1': [(0.25, 0), (0.75, 0), (1, 0.25), (1, 0.75), (0.75, 1), (0.25, 1), (0, 0.75), (0, 0.25)],
    'fe': [(0.25, 0.5), (0.75, 0.5), (0.5, 0.75), (0.5, 0.25)],
    'fe!': [(-0.2, 0.5), (1.2, 0.5), (0.5, 1.22), (0.5, -0.2)]
}

# Interior point classes require degree >= MIN_DEGREE_INTERIOR unless a continuation edge is present
MIN_DEGREE_INTERIOR = 3
RANKS = [1, 2, 3]
interior_point_classes = {'F', 'vf', 'fe'}
interior_positions = set(
    p for cls in interior_point_classes for p in dot_positions[cls]
)


def _vf_positions(t):
    """vf positions at parameter t along V→F (0 = at V corner, 1 = at F centre)."""
    fx, fy = dot_positions['F'][0]
    return [(vx + t * (fx - vx), vy + t * (fy - vy)) for vx, vy in dot_positions['V']]


def _vf_outside_positions(vf_pts):
    """Compute vf! positions to match given vf positions.
    Each vf! point tracks the x or y coordinate of the corresponding vf vertex,
    projected outside the cell boundary."""
    NEAR, FAR = -0.2, 1.2
    x0, y0 = vf_pts[0]
    x1, y1 = vf_pts[1]
    x2, y2 = vf_pts[2]
    x3, y3 = vf_pts[3]
    return [
        (x0, NEAR),   # below vf[0]
        (x1, NEAR),   # below vf[1]
        (FAR,  y1),   # right of vf[1]
        (FAR,  y2),   # right of vf[2]
        (x2, FAR),    # above vf[2]
        (x3, FAR),    # above vf[3]
        (NEAR, y3),   # left of vf[3]
        (NEAR, y0),   # left of vf[0]
    ]


def _fe_positions(t):
    """fe positions at parameter t (0.5 = default, 0 = collapsed to F, 1 = at E midpoint).
    Parameterises each fe[i] along the F→fe[i]→E line, preserving index order."""
    fx, fy = dot_positions['F'][0]
    # fe[i] sits at t=0.5 by default; scaling by 2t keeps t=0.5 → default positions.
    return [(fx + 2 * t * (px - fx), fy + 2 * t * (py - fy))
            for px, py in dot_positions['fe']]


def _effective_interior(vf_pts=None, fe_pts=None):
    """Interior position set, updated for non-default vf/fe placements."""
    return (
        set(map(tuple, dot_positions['F'])) |
        set(map(tuple, vf_pts if vf_pts is not None else dot_positions['vf'])) |
        set(map(tuple, fe_pts if fe_pts is not None else dot_positions['fe']))
    )


# Candidate placements tried for both vf and fe when checking for crossings.
# vf slides V→F (t=0 at corner, t=1 at centre); fe slides F→E (t=0 at centre, t=1 at edge).
# Default midpoint (1/2) is tried first so it is used for rendering whenever it is crossing-free.
_T_VALUES = [1/2, 1/3, 2/3]


atoms = {
    ('E', 'E'): [(0, [1]), (1, [2]), (2, [3]), (3, [0])],
    ('E', 'F'): [(0, [0]), (1, [0]), (2, [0]), (3, [0])],
    ('E', 'V'): [(0, [0, 1]), (1, [1, 2]), (2, [2, 3]), (3, [3, 0])],
    ('E', 've'): [(0, [0, 1]), (1, [2, 3]), (2, [4, 5]), (3, [6, 7])],
    ('E', 'vf'): [(0, [0, 1]), (1, [1, 2]), (2, [2, 3]), (3, [3, 0])],
    ('E', 'fe'): [(0, [3]), (1, [1]), (2, [2]), (3, [0])],
    ('F', 'F!'): [(0, [0, 1, 2, 3])],
    ('F', 'V'): [(0, [0, 1, 2, 3])],
    ('F', 've'): [(0, [0, 1, 2, 3, 4, 5, 6, 7])],
    ('F', 'vf'): [(0, [0, 1, 2, 3])],
    ('F', 'fe'): [(0, [0, 1, 2, 3])],
    ('V', 'V'): [(0, [1, 3]), (1, [0, 2]), (2, [1, 3]), (3, [0, 2])],
    ('V', 've'): [(0, [0, 7]), (1, [1, 2]), (2, [3, 4]), (3, [5, 6])],
    ('V', 'vf'): [(0, [0]), (1, [1]), (2, [2]), (3, [3])],
    ('fe', 'V'): [(0, [3]), (0, [0]), (1, [2]), (1, [1]), (2, [2]), (2, [3]), (3, [0]), (3, [1])],
    # ve-ve: split into edge-adjacent (ve0) and corner-adjacent (ve1)
    ('ve0', 've0'): [(0, [1]), (1, [0]), (2, [3]), (3, [2]), (4, [5]), (5, [4]), (6, [7]), (7, [6])],
    ('ve1', 've1'): [(0, [7]), (1, [2]), (2, [1]), (3, [4]), (4, [3]), (5, [6]), (6, [5]), (7, [0])],
    ('ve', 'vf'): [(0, [0]), (1, [1]), (2, [1]), (3, [2]), (4, [2]), (5, [3]), (6, [3]), (7, [0])],
    ('fe', 've'): [(0, [6, 7]), (1, [2, 3]), (2, [4, 5]), (3, [0, 1])],
    ('vf', 'vf'): [(0, [1, 3]), (1, [0, 2]), (2, [1, 3]), (3, [0, 2])],
    ('vf', 'vf!'): [(0, [7, 0]), (1, [1, 2]), (2, [3, 4]), (3, [5, 6])],
    ('fe', 'vf'): [(0, [3, 0]), (1, [1, 2]), (2, [2, 3]), (3, [0, 1])],
    ('fe', 'fe'): [(0, [3]), (1, [2]), (2, [0]), (3, [1])],
    ('fe', 'fe!'): [(0, [0]), (1, [1]), (2, [2]), (3, [3])]
}

EPSILON = 1e-9


@dataclass(frozen=True)
class Segment:
    a: tuple
    b: tuple

    def __post_init__(self):
        # Canonical order so (a,b) == (b,a)
        if self.a > self.b:
            a, b = self.b, self.a
            object.__setattr__(self, 'a', a)
            object.__setattr__(self, 'b', b)


def cross2d(ox, oy, ax, ay, bx, by):
    return (ax - ox) * (by - oy) - (ay - oy) * (bx - ox)


def segments_intersect(s1: Segment, s2: Segment) -> bool:
    """Returns True if s1 and s2 intersect in their interiors (not at shared endpoints)."""
    p1, p2 = s1.a, s1.b
    p3, p4 = s2.a, s2.b

    # Shared endpoint — touching is allowed
    shared = {p1, p2} & {p3, p4}
    if len(shared) == 2:
        return True  # duplicate segment
    # For shared endpoint: non-collinear touching is fine, but collinear overlap is not.
    # Fall through — the parametric check returns False for non-collinear shared endpoints
    # (t or u lands on 0 or 1, outside the open interval), and the collinear path
    # correctly detects one segment containing the other.

    d1x, d1y = p2[0] - p1[0], p2[1] - p1[1]
    d2x, d2y = p4[0] - p3[0], p4[1] - p3[1]

    denom = d1x * d2y - d1y * d2x

    if abs(denom) < EPSILON:
        # Parallel — check if actually collinear before testing overlap
        cross = (p3[0] - p1[0]) * d1y - (p3[1] - p1[1]) * d1x
        if abs(cross) > EPSILON:
            return False  # parallel but not collinear
        return collinear_overlap(s1, s2)

    t = ((p3[0] - p1[0]) * d2y - (p3[1] - p1[1]) * d2x) / denom
    u = ((p3[0] - p1[0]) * d1y - (p3[1] - p1[1]) * d1x) / denom

    t_in = EPSILON < t < 1 - EPSILON
    u_in = EPSILON < u < 1 - EPSILON
    # Also flag T-intersections: endpoint of one segment lying strictly inside the other.
    # (Proper shared-endpoint touching has t≈0/1 AND u≈0/1, so neither _in is True.)
    u_end = abs(u) < EPSILON or abs(u - 1) < EPSILON
    t_end = abs(t) < EPSILON or abs(t - 1) < EPSILON
    return (t_in and (u_in or u_end)) or (u_in and t_end)


def collinear_overlap(s1: Segment, s2: Segment) -> bool:
    """Returns True if two collinear segments overlap (share more than a point)."""
    # Project onto whichever axis has more spread
    dx = s1.b[0] - s1.a[0]
    dy = s1.b[1] - s1.a[1]

    if abs(dx) > abs(dy):
        a1, b1 = s1.a[0], s1.b[0]
        a2, b2 = s2.a[0], s2.b[0]
    else:
        a1, b1 = s1.a[1], s1.b[1]
        a2, b2 = s2.a[1], s2.b[1]

    lo1, hi1 = min(a1, b1), max(a1, b1)
    lo2, hi2 = min(a2, b2), max(a2, b2)

    overlap = min(hi1, hi2) - max(lo1, lo2)
    return overlap > EPSILON


def _base_class(cls: str) -> str:
    """Normalise a point class name to its base class for rank calculation."""
    cls = cls.rstrip('!')
    for suffix in ('_e', '_c'):
        if cls.endswith(suffix):
            return cls[:-len(suffix)]
    return cls


def combination_rank(combo) -> int:
    """Number of distinct base point classes in an atom combination."""
    return len({_base_class(cls) for atom in combo for cls in atom})


def connectivity_check(segments: list) -> bool:
    """Returns True if the segment graph is connected (single component)."""
    if not segments:
        return False
    adj = {}
    for seg in segments:
        adj.setdefault(seg.a, set()).add(seg.b)
        adj.setdefault(seg.b, set()).add(seg.a)
    nodes = list(adj)
    visited = {nodes[0]}
    queue = [nodes[0]]
    while queue:
        node = queue.pop()
        for nb in adj[node]:
            if nb not in visited:
                visited.add(nb)
                queue.append(nb)
    return visited == set(nodes)


_boundary_point_positions = (
    set(map(tuple, dot_positions['V'])) |
    set(map(tuple, dot_positions['E'])) |
    set(map(tuple, dot_positions['ve'])) |
    set(map(tuple, dot_positions['ve0'])) |
    set(map(tuple, dot_positions['ve1']))
)
_outside_positions = (
    set(map(tuple, dot_positions['F!'])) |
    set(map(tuple, dot_positions['vf!'])) |
    set(map(tuple, dot_positions['fe!']))
)


def adjacent_face_check(segments: list, points: list) -> bool:
    """Returns True if the operator connects to adjacent faces — either via
    boundary points (V, E, ve) or via continuation segments to outside (!) points."""
    if any(p in _boundary_point_positions for p in set(points)):
        return True
    if any(seg.a in _outside_positions or seg.b in _outside_positions for seg in segments):
        return True
    return False


def has_crossing_or_duplicate(segments: list) -> bool:
    """Returns True if any two segments cross or are duplicates."""
    for i in range(len(segments)):
        for j in range(i + 1, len(segments)):
            if segments_intersect(segments[i], segments[j]):
                return True
    return False


def min_interior_degree(segments: list, points: list, vf_pts=None, fe_pts=None) -> int:
    """Returns the minimum degree of any interior point (F, vf, fe), or MIN_DEGREE_INTERIOR if none present."""
    degree = {}
    for seg in segments:
        degree[seg.a] = degree.get(seg.a, 0) + 1
        degree[seg.b] = degree.get(seg.b, 0) + 1
    eff_interior = _effective_interior(vf_pts, fe_pts)
    interior_pts = [p for p in set(points) if p in eff_interior]
    if not interior_pts:
        return MIN_DEGREE_INTERIOR
    return min(degree.get(p, 0) for p in interior_pts)


def degree_check(segments: list, points: list, vf_pts=None, fe_pts=None) -> bool:
    """Returns True if all interior vertices (F, vf, fe) have degree >= MIN_DEGREE_INTERIOR."""
    return min_interior_degree(segments, points, vf_pts, fe_pts) >= MIN_DEGREE_INTERIOR


# Corner positions — shared by 4 cells, too complex to analyse locally
corner_positions = set(map(tuple, dot_positions['V']))


def _boundary_edge(pt):
    """Returns the edge index (0=bottom,1=right,2=top,3=left) for a non-corner boundary point,
    or None if the point is a corner, interior, or outside the cell."""
    if pt in corner_positions:
        return None
    x, y = pt
    if abs(y) < EPSILON:       return 0
    if abs(x - 1) < EPSILON:   return 1
    if abs(y - 1) < EPSILON:   return 2
    if abs(x) < EPSILON:       return 3
    return None


def _reflect(pt, edge_idx):
    x, y = pt
    if edge_idx == 0: return (x, -y)
    if edge_idx == 1: return (2 - x, y)
    if edge_idx == 2: return (x, 2 - y)
    if edge_idx == 3: return (-x, y)


def min_boundary_degree(segments: list, points: list) -> int:
    """Returns the minimum global degree of any non-corner boundary point (E, ve)
    after simulating the adjacent cell's contribution via reflection.
    Returns MIN_DEGREE_INTERIOR if no such points exist."""
    degree = {}
    incident = {}
    for seg in segments:
        for pt, other in ((seg.a, seg.b), (seg.b, seg.a)):
            degree[pt] = degree.get(pt, 0) + 1
            incident.setdefault(pt, set()).add(other)

    min_deg = MIN_DEGREE_INTERIOR
    for pt in set(points):
        local_deg = degree.get(pt, 0)
        if local_deg >= MIN_DEGREE_INTERIOR:
            continue
        edge_idx = _boundary_edge(pt)
        if edge_idx is None:
            continue
        neighbors = frozenset(incident.get(pt, set()))
        reflected = frozenset(_reflect(n, edge_idx) for n in neighbors)
        new_connections = reflected - neighbors - {pt}
        global_deg = local_deg + len(new_connections)
        min_deg = min(min_deg, global_deg)
    return min_deg


def create_data(line_config, vf_pts=None, fe_pts=None):
    pos = dict(dot_positions)
    if vf_pts is not None:
        pos['vf'] = vf_pts
        pos['vf!'] = _vf_outside_positions(vf_pts)
    if fe_pts is not None:
        pos['fe'] = fe_pts
    line_config = line_config.split(',')
    points = []
    segments = []

    for line in line_config:
        group_a, group_b = line.split('-')
        if group_a.upper() > group_b.upper():
            start_group, end_group = group_b, group_a
        else:
            start_group, end_group = group_a, group_b
        key = (start_group, end_group)

        start_points = pos[start_group]
        end_points = pos[end_group]
        all_indices = atoms[key]
        omit_end = key[1].endswith("!")

        points += list(start_points)
        if not omit_end:
            points += list(end_points)

        atom_segments = []
        for connection_tuple in all_indices:
            start = start_points[connection_tuple[0]]
            for i in connection_tuple[1]:
                end = end_points[i]
                atom_segments.append(Segment(start, end))
        # Deduplicate within atom (atom definitions list each edge from both ends)
        segments.extend(dict.fromkeys(atom_segments))

    return segments, points


def _build_combo(combo, vf_pts=None, fe_pts=None):
    """Build (segments, dots, pair_indices) for a combo with given vf/fe positions."""
    segs, dots, indices = [], [], []
    for idx, atom in enumerate(combo):
        s, d = create_data('-'.join(atom), vf_pts=vf_pts, fe_pts=fe_pts)
        segs += s
        dots += d
        indices += [idx] * len(s)
    return segs, dots, indices


def _find_valid_placement(combo):
    """Try all (vf_t, fe_t) combinations; return (segments, dots, indices, vf_pts, fe_pts)
    for the first crossing-free placement, or (None, None, None, None, None) if all fail."""
    for vf_t in _T_VALUES:
        for fe_t in _T_VALUES:
            vf_pts = _vf_positions(vf_t)
            fe_pts = _fe_positions(fe_t)
            segs, dots, indices = _build_combo(combo, vf_pts, fe_pts)
            if not has_crossing_or_duplicate(segs):
                return segs, dots, indices, vf_pts, fe_pts
    return None, None, None, None, None


RENDER_MODE = 'final'  # 'debug' or 'final'

LINE_WIDTH = 0.04
DOT_RADIUS = 0.075
DOT_OUTLINE_WIDTH = 0.015

colors = ['red', 'green', 'blue']
widths = ['0.1', '0.05', '0.025']
opacities = ['0.3', '0.5', '0.7']


def _is_outside_pt(pt):
    """Returns True if pt is outside the [0,1] cell (adjacent-face stub)."""
    x, y = pt
    return x < -0.05 or x > 1.05 or y < -0.05 or y > 1.05


def _tapering_line_svg(interior, outside, stroke_width, taper=0.25):
    """Render a segment as a normal-width line that tapers to a point over the last `taper` fraction."""
    x1, y1 = interior
    x2, y2 = outside
    dx, dy = x2 - x1, y2 - y1
    length = math.sqrt(dx * dx + dy * dy)
    if length < 1e-9:
        return ''
    px, py = -dy / length, dx / length  # perpendicular unit vector
    hw = stroke_width / 2
    # Point where taper begins
    tx = x1 + (1 - taper) * dx
    ty = y1 + (1 - taper) * dy
    # Pentagon: rectangle from interior to taper point, then triangle to outside tip
    pts = [
        (x1 + px * hw, y1 + py * hw),
        (tx + px * hw, ty + py * hw),
        (x2, y2),
        (tx - px * hw, ty - py * hw),
        (x1 - px * hw, y1 - py * hw),
    ]
    pts_str = ' '.join(f'{x:.4f},{y:.4f}' for x, y in pts)
    return f'<polygon points="{pts_str}" fill="black" stroke="none"/>'


def create_svg(all_segs, all_pts, indices, mode=RENDER_MODE):
    unique_dots = list(set(all_pts))
    padding = 0.25
    page_size = 200
    viewbox_size = 1 + 2 * padding
    svg_header = f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="-{padding} -{padding} {viewbox_size} {viewbox_size}" width="{page_size}" height="{page_size}">'
    svg_content = ""

    if mode == 'final':
        svg_content += f'<rect x="0" y="0" width="1" height="1" fill="none" stroke="#bbb" stroke-width="0.048" stroke-dasharray="0.04 0.025"/>'
        for seg in all_segs:
            a_out = _is_outside_pt(seg.a)
            b_out = _is_outside_pt(seg.b)
            if a_out or b_out:
                interior = seg.b if a_out else seg.a
                outside = seg.a if a_out else seg.b
                svg_content += _tapering_line_svg(interior, outside, LINE_WIDTH)
            else:
                svg_content += f'<line x1="{seg.a[0]}" y1="{seg.a[1]}" x2="{seg.b[0]}" y2="{seg.b[1]}" stroke="black" stroke-width="{LINE_WIDTH}"/>'
        for x, y in unique_dots:
            svg_content += f'<circle cx="{x}" cy="{y}" r="{DOT_RADIUS}" fill="red" stroke="black" stroke-width="{DOT_OUTLINE_WIDTH}"/>'
    else:
        for count, seg in enumerate(all_segs):
            idx = indices[count]
            svg_content += f'<line x1="{seg.a[0]}" y1="{seg.a[1]}" x2="{seg.b[0]}" y2="{seg.b[1]}" stroke="{colors[idx]}" opacity="{opacities[idx]}" stroke-width="{widths[idx]}"/>'
        for x, y in unique_dots:
            svg_content += f'<circle cx="{x}" cy="{y}" r="0.05" fill="purple" stroke="black" stroke-width="0.01"/>'

    svg_footer = '</svg>'
    return f"{svg_header}{svg_content}{svg_footer}"


def svg_to_png(svg_path: str, png_path: str):
    """Rasterize an SVG file to PNG using svglib + pycairo."""
    renderPM.drawToFile(svg2rlg(svg_path), png_path, fmt='PNG')


def open_svg(filename):
    os_name = platform.system()
    if os_name == 'Darwin':
        subprocess.run(['open', filename], check=True)
    elif os_name == 'Windows':
        subprocess.run(['start', filename], shell=True, check=True)


OPERATOR_METADATA = {
    'E-E':                         {'name': 'Ambo'},
    'F-F!':                        {'name': 'Dual'},
    'V-V':                         {'name': 'Seed',               'gc_symbol': 'u1 / o1'},
    'fe-fe,fe-fe!':                {'name': 'Zip'},
    've0-ve0,ve1-ve1':             {'name': 'Truncate'},
    'vf-vf,vf-vf!':                {'name': 'Expand'},
    'E-E,E-F':                     {'conway_symbol': 'dc/ud',      'dual': 'Chamfer'},
    'F-V':                         {'name': 'Join',                'gc_symbol': 'o1,1'},
    'E-E,E-V':                     {'name': 'Subdivide',           'gc_symbol': 'u2'},
    'F-F!,F-V':                    {'name': 'Needle',              'gc_symbol': 'u1,1'},
    'F-V,V-V':                     {'name': 'Kis'},
    'E-E,E-fe,fe-fe':              {'conway_symbol': 'dl'},
    'E-vf,vf-vf':                  {'dual': 'Stake?'},
    'E-vf,vf-vf!':                 {'dual': 'Stake?',              'concave': '(Y)'},
    'F-ve,ve0-ve0':                {'concave': 'Y'},
    'F-vf,vf-vf!':                 {'conway_symbol': 'dqd'},
    'fe-V,fe-fe':                  {'name': 'Join-Lace'},
    'fe-V,fe-fe,fe-fe!':           {'name': 'Opposite-Lace'},
    'V-V,fe-V,fe-fe':              {'name': 'Lace'},
    'V-ve,ve0-ve0,ve1-ve1':        {'conway_symbol': 'dld'},
    'V-vf,vf-vf':                  {'name': 'Chamfer'},
    'V-V,V-vf,vf-vf':              {'name': 'Loft'},
    've0-ve0,fe-ve,fe-fe':         {'concave': 'Y'},
    've1-ve1,fe-ve,fe-fe!':        {'concave': 'Y'},
    'E-F,E-V':                     {'name': 'Ortho',               'gc_symbol': 'o2'},
    'E-F,E-V,F-V':                 {'name': 'Meta'},
    'E-F,E-ve,F-ve':               {'name': '(Join-Edge-Medial)',  'concave': 'Y'},
    'E-V,E-fe,fe-fe':              {'name': 'Quinto'},
    'E-V,E-fe,fe-V':               {'concave': 'Y'},
    'E-vf,V-vf':                   {'concave': '(Y)'},
    'E-vf,V-vf,vf-vf!':            {'concave': '(Y)'},
    'E-V,E-vf,V-vf':               {'concave': '(Y)'},
    'E-V,E-vf,vf-vf':              {'name': '(Squall)',            'concave': '(Y)'},
    'E-ve,E-fe,fe-ve,fe-fe':       {'concave': 'Y'},
    'E-ve,E-fe,ve1-ve1,fe-ve':     {'concave': 'Y'},
    'E-vf,E-fe,fe-vf':             {'concave': 'Y'},
    'F-fe,fe-V':                   {'name': 'Join-Stake'},
    'F-fe,fe-V,fe-fe!':            {'name': 'Opposite-Stake'},
    'F-fe,V-V,fe-V':               {'name': 'Stake'},
    'F-V,F-fe,fe-V':               {'name': 'Join-Kis-Kis'},
}


def _notation_to_file_class(notation):
    classes = set()
    for atom in notation.split(','):
        for part in atom.split('-'):
            classes.add(_base_class(part))
    return '.'.join(sorted(classes))


def _notation_to_display_class(notation):
    _VE_NORM = {'ve0': 've', 've1': 've'}
    classes = set()
    for atom in notation.split(','):
        for part in atom.split('-'):
            base = _base_class(part)
            classes.add(_VE_NORM.get(base, base))
    return '.'.join(sorted(classes))


def generate_sheet(quality='good', output_path=None):
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    if output_path is None:
        output_path = f'operators_{quality}.xlsx'

    wb = Workbook()
    ws = wb.active
    ws.title = 'Operators'

    headers = ['Rank', 'Class', '# Atoms', 'Diagram', 'Notation',
               'Conway/Hart Name', 'Conway Symbol', 'Goldberg-Coxeter Symbol', 'Dual', 'Concave Faces?']
    ws.append(headers)
    for col, _ in enumerate(headers, 1):
        ws.cell(1, col).font = Font(bold=True)
        ws.cell(1, col).alignment = Alignment(horizontal='center', vertical='center')

    col_widths_chars = [8, 12, 10, 18, 32, 22, 16, 26, 16, 14]
    for i, w in enumerate(col_widths_chars, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    IMG_SIZE = 90
    IMG_ROW_HEIGHT = 70

    csv_path = f'operators_{quality}.csv'
    operators = []
    with open(csv_path, newline='') as f:
        for row in csv.DictReader(f):
            operators.append(row)

    for i, op in enumerate(operators):
        rank = op['rank']
        notation = op['operator']
        file_class = _notation_to_file_class(notation)
        display_class = _notation_to_display_class(notation)
        n_atoms = len(notation.split(','))
        row_num = i + 2

        meta = OPERATOR_METADATA.get(notation, {})
        ws.append([
            int(rank), display_class, n_atoms, '', notation,
            meta.get('name', ''), meta.get('conway_symbol', ''),
            meta.get('gc_symbol', ''), meta.get('dual', ''), meta.get('concave', '')
        ])
        for col in range(1, len(headers) + 1):
            ws.cell(row_num, col).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row_num, 5).alignment = Alignment(horizontal='left', vertical='center')

        png_path = f'Rank {rank}/pngs/{quality}/{file_class}/{notation}.png'
        if os.path.exists(png_path):
            img = XLImage(png_path)
            img.width = IMG_SIZE
            img.height = IMG_SIZE
            ws.add_image(img, f'D{row_num}')
            ws.row_dimensions[row_num].height = IMG_ROW_HEIGHT
            print(f'  [{i+1}/{len(operators)}] {notation}')
        else:
            print(f'  [{i+1}/{len(operators)}] missing PNG: {png_path}')

    ws.freeze_panes = 'A2'
    wb.save(output_path)
    print(f'Saved: {output_path}')


def build_crossing_matrix(output_path='crossing_matrix.csv'):
    """Write a CSV table: rows and columns are atoms, cells mark pairs that are
    crossing_or_duplicate for ALL tried vf placements (X = always crosses, blank = avoidable)."""
    atom_keys = list(atoms.keys())
    atom_labels = ['-'.join(k) for k in atom_keys]

    with open(output_path, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow([''] + atom_labels)
        for key_a in atom_keys:
            row = ['-'.join(key_a)]
            for key_b in atom_keys:
                combo = (key_a, key_b)
                segs, _, _, _, _ = _find_valid_placement(combo)
                row.append('' if segs is not None else 'X')
            writer.writerow(row)


if __name__ == '__main__':
    if len(sys.argv) > 1 and sys.argv[1] == 'generate-sheet':
        q = sys.argv[2] if len(sys.argv) > 2 else 'good'
        generate_sheet(quality=q)
        sys.exit(0)

    if len(sys.argv) > 1 and sys.argv[1] == 'crossing-matrix':
        output = sys.argv[2] if len(sys.argv) > 2 else 'crossing_matrix.csv'
        build_crossing_matrix(output)
        sys.exit(0)

    # For each target rank, the max useful atoms is the largest number of atoms
    # whose base classes are all subsets of some rank-R base class set.
    def max_atoms_for_rank(rank):
        from itertools import combinations as _combinations
        base_classes = list({_base_class(cls) for a in atoms for cls in a if not cls.endswith('!')})
        best = 0
        for class_set in _combinations(base_classes, rank):
            class_set = frozenset(class_set)
            count = sum(1 for a in atoms if {_base_class(c) for c in a} <= class_set)
            best = max(best, count)
        return best

    ALL_QUALITIES = ['good', 'degree2', 'no_adjacent_face', 'bad_connectivity', 'low_degree', 'crossing_or_duplicate']

    seen_combos = set()
    csv_files = {q: open(f'operators_{q}.csv', 'w', newline='') for q in ALL_QUALITIES}
    csv_writers = {q: csv.writer(f) for q, f in csv_files.items()}
    for writer in csv_writers.values():
        writer.writerow(['rank', 'operator'])

    try:
        for target_rank in RANKS:
            max_atoms = max_atoms_for_rank(target_rank)
            for n_atoms in range(1, max_atoms + 1):
                for combo in itertools.combinations(atoms.keys(), r=n_atoms):
                    if combination_rank(combo) != target_rank:
                        continue
                    if combo in seen_combos:
                        continue
                    seen_combos.add(combo)

                    pair = combo
                    readable_name = str(pair).replace(' ', '').replace('\',\'', '-').replace('(', '').replace(')', '').replace('\'', '')
                    if readable_name.endswith(','): readable_name = readable_name[:-1]

                    all_segments, all_dots, pair_indices, vf_pts, fe_pts = _find_valid_placement(pair)
                    if all_segments is None:
                        # All placements produced crossings — use defaults for rendering
                        all_segments, all_dots, pair_indices = _build_combo(pair)
                        vf_pts = fe_pts = None
                        quality = 'crossing_or_duplicate'
                    else:
                        min_deg = min(min_interior_degree(all_segments, all_dots, vf_pts, fe_pts),
                                     min_boundary_degree(all_segments, all_dots))
                        if min_deg <= 1:
                            quality = 'low_degree'
                        elif not connectivity_check(all_segments):
                            quality = 'bad_connectivity'
                        elif not adjacent_face_check(all_segments, all_dots):
                            quality = 'no_adjacent_face'
                        elif min_deg == 2:
                            quality = 'degree2'
                        else:
                            quality = 'good'

                    csv_writers[quality].writerow([target_rank, readable_name])

                    class_subdir = '.'.join(sorted({_base_class(cls) for atom in pair for cls in atom}))
                    subpath = f'{quality}/{class_subdir}/{readable_name}'
                    rank_subdir = f'Rank {target_rank}'
                    os.makedirs(f'{rank_subdir}/svgs/{quality}/{class_subdir}', exist_ok=True)
                    os.makedirs(f'{rank_subdir}/pngs/{quality}/{class_subdir}', exist_ok=True)
                    svg = create_svg(all_segments, all_dots, pair_indices)
                    with open(f'{rank_subdir}/svgs/{subpath}.svg', 'w') as file:
                        file.write(svg)
                    svg_to_png(f'{rank_subdir}/svgs/{subpath}.svg', f'{rank_subdir}/pngs/{subpath}.png')
    finally:
        for f in csv_files.values():
            f.close()